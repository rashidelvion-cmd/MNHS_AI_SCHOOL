"""
Parser for the official SF1-SHS registers.

Two formats are supported through declarative profiles (see profiles.py):
  * 2026 official template   (sheet "SHSF-1",   .xlsx)
  * legacy 2018.2.1.1        (sheet "school_form_1_shs_ver2018*", .xls/.xlsx)

The format is auto-detected by sheet name / header signature, so callers
use parse_sf1_workbook() unchanged regardless of the file version.

This module only reads the workbook and validates rows. It never touches
the database and never modifies the uploaded file.
"""

from __future__ import annotations

import re
from collections import Counter

import openpyxl

from ..base import (
    ParsedRow,
    ParseResult,
    clean_text,
    is_placeholder_lrn,
    lrn_looks_official,
    normalize_date,
    normalize_lrn,
    normalize_sex,
)
from .profiles import detect_profile
from .xls_loader import is_xls, load_xls_as_openpyxl

OFFICIAL_REMARKS_CODES = {"T/O", "T/I", "CCT", "B/A", "LWE", "ACL"}


def _row_text(worksheet, row_number, max_col):
    """All non-empty strings on a header row, joined for display."""
    values = []
    for col in range(1, max_col + 1):
        value = worksheet.cell(row=row_number, column=col).value
        text = clean_text(value)
        if text:
            values.append(text)
    return "  |  ".join(values)


def parse_sf1_workbook(path):
    """
    Parse an SF1-SHS workbook (any supported format) into a ParseResult.

    File-level errors abort with ParseResult.file_errors set. Row-level
    problems land on the individual ParsedRow as errors/warnings.

    Formula robustness (2026 .xlsx): files produced by the school's Excel
    automation may hold formulas with no cached values. A second
    formula-view workbook resolves literal formulas (="TEXT") and detects
    truly uncached cells so an actionable error is reported.
    """
    result = ParseResult()

    # ---- open the workbook (with a .xls shim when needed) ----
    formula_workbook = None
    try:
        if is_xls(path):
            workbook = load_xls_as_openpyxl(path)
            # .xls has no separate formula layer; values are already resolved.
        else:
            workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
            formula_workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception:
        result.file_errors.append(
            "The file could not be opened as an Excel workbook. Please upload "
            "the official SF1-SHS file (.xlsx or .xls)."
        )
        return result

    # ---- detect the format profile ----
    profile, worksheet = detect_profile(workbook)
    if profile is None:
        result.file_errors.append(
            "This workbook does not match any known SF1-SHS layout "
            "(the LRN/NAME header row was not found where expected). "
            "Please upload an official SF1 School Register for SHS."
        )
        return result

    formula_sheet = None
    if formula_workbook is not None:
        # match the same sheet in the formula view
        if worksheet.title in formula_workbook.sheetnames:
            formula_sheet = formula_workbook[worksheet.title]

    result.header = {
        "profile": profile.key,
        "school_row": _row_text(worksheet, profile.header_rows["school_name_row"], profile.header_max_col),
        "context_row": _row_text(worksheet, profile.header_rows["context_row"], profile.header_max_col),
        "section_row": _row_text(worksheet, profile.header_rows["section_row"], profile.header_max_col),
    }

    uncached_formula_cells = 0

    for block in profile.blocks:
        counter = 0
        for sheet_row in range(block.first_row, block.last_row + 1):
            counter += 1
            raw = {
                field_name: worksheet[f"{col}{sheet_row}"].value
                for field_name, col in profile.columns.items()
            }
            if formula_sheet is not None:
                raw, uncached = _resolve_formula_cells(raw, formula_sheet, sheet_row, profile)
                uncached_formula_cells += uncached

            lrn_raw = raw["lrn"]
            name_raw = clean_text(raw["name"])

            # Skip pre-printed placeholder / empty template lines.
            if profile.has_placeholder_counters:
                if is_placeholder_lrn(lrn_raw, counter) and not name_raw:
                    continue
            else:
                # Legacy files have no counters: an empty row = blank LRN+name.
                if normalize_lrn(lrn_raw) == "" and not name_raw:
                    continue

            row = ParsedRow(sheet_row=sheet_row, block=block.label)
            _parse_learner_row(row, raw, counter, profile)
            result.rows.append(row)

    _flag_duplicate_lrns_in_file(result)

    if not result.rows:
        if uncached_formula_cells:
            result.file_errors.append(
                "The learner rows in this file are filled by Excel formulas "
                "whose calculated values were not saved (the file was "
                "generated by a program and never re-saved in Excel). "
                "Please open the file in Microsoft Excel, press Save once, "
                "and upload it again — or upload a copy where the learner "
                "data is pasted as values."
            )
        else:
            result.file_errors.append(
                "No learner rows were found in the file — the register "
                "appears to be empty."
            )

    return result


_LITERAL_TEXT_FORMULA = re.compile(r'^\s*=\s*"(.*)"\s*$')
_LITERAL_NUMBER_FORMULA = re.compile(r"^\s*=\s*(-?\d+(?:\.\d+)?)\s*$")


def _resolve_formula_cells(raw, formula_sheet, sheet_row, profile):
    """
    For cells that read as None under data_only, consult the formula view:
    resolve literal formulas (="TEXT", =12345) to their value; count real
    formulas with no cached value so the caller can report them.
    """
    uncached = 0
    resolved = dict(raw)
    for field_name, col in profile.columns.items():
        if resolved[field_name] is not None:
            continue
        formula_value = formula_sheet[f"{col}{sheet_row}"].value
        if not isinstance(formula_value, str) or not formula_value.startswith("="):
            continue
        text_match = _LITERAL_TEXT_FORMULA.match(formula_value)
        if text_match:
            resolved[field_name] = text_match.group(1)
            continue
        number_match = _LITERAL_NUMBER_FORMULA.match(formula_value)
        if number_match:
            resolved[field_name] = number_match.group(1)
            continue
        uncached += 1
    return resolved, uncached


def _parse_learner_row(row, raw, counter, profile):
    lrn = normalize_lrn(raw["lrn"])
    name_raw = clean_text(raw["name"])

    if profile.has_placeholder_counters and is_placeholder_lrn(raw["lrn"], counter):
        lrn = ""

    if not lrn and name_raw:
        row.errors.append("LRN is missing for this learner.")
    if lrn and not name_raw:
        row.errors.append("Name is missing for this learner.")

    if lrn and not lrn_looks_official(lrn):
        row.warnings.append(
            f"LRN '{lrn}' is not the standard 12-digit format — it will be "
            "imported as written."
        )

    last, first, extension, middle, name_warning = profile.split_name(name_raw)
    if name_warning:
        row.warnings.append(name_warning)

    sex, sex_warning = normalize_sex(raw["sex"], block_label=row.block)
    if sex_warning:
        row.warnings.append(sex_warning)

    birth_date, date_error = normalize_date(raw["birth_date"])
    if date_error:
        row.errors.append(date_error)
    elif birth_date is None:
        row.errors.append("Birthdate is missing (required).")

    remarks = clean_text(raw["remarks_code"])
    if remarks:
        codes = [c.strip().upper() for c in remarks.replace(";", ",").split(",") if c.strip()]
        unknown = [c for c in codes if c not in OFFICIAL_REMARKS_CODES]
        if unknown:
            row.warnings.append(
                "Remarks contain codes outside the official legend "
                f"({', '.join(unknown)}) — kept as written."
            )

    row.data = {
        "lrn": lrn,
        "last_name": last,
        "first_name": first,
        "name_extension": extension,
        "middle_name": middle,
        "gender": sex,
        "birth_date": birth_date,
        "religious_affiliation": clean_text(raw["religious_affiliation"]),
        "house_street": clean_text(raw["house_street"]),
        "barangay": clean_text(raw["barangay"]),
        "municipality": clean_text(raw["municipality"]),
        "province": clean_text(raw["province"]),
        "father_name": clean_text(raw["father_name"]),
        "mother_maiden_name": clean_text(raw["mother_maiden_name"]),
        "guardian_name": clean_text(raw["guardian_name"]),
        "guardian_relationship": clean_text(raw["guardian_relationship"]),
        "contact_number": clean_text(raw["contact_number"]),
        "remarks_code": remarks,
    }


def _flag_duplicate_lrns_in_file(result):
    counts = Counter(r.data.get("lrn") for r in result.rows if r.data.get("lrn"))
    duplicates = {lrn for lrn, count in counts.items() if count > 1}
    if not duplicates:
        return
    for row in result.rows:
        if row.data.get("lrn") in duplicates:
            row.errors.append(
                f"LRN {row.data['lrn']} appears more than once in this file — "
                "all affected rows were excluded. Please fix the file and "
                "re-upload."
            )
