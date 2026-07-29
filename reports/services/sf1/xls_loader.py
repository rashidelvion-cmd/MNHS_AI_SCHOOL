"""
Legacy .xls (BIFF) support for SF1 import.

openpyxl cannot read the old binary .xls format, so a .xls upload is read
with xlrd and copied cell-for-cell into an in-memory openpyxl workbook.
Downstream parsing then works identically for .xls and .xlsx, so no
version-specific code is needed beyond this shim.
"""

from __future__ import annotations

import datetime

import openpyxl


def is_xls(path):
    return str(path).lower().endswith(".xls")


def load_xls_as_openpyxl(path):
    """
    Read a .xls workbook with xlrd and return an equivalent openpyxl
    Workbook (values only). Dates are converted to datetime; other values
    are copied verbatim. Raises on unreadable files (caller handles it).
    """
    import xlrd  # imported lazily so .xlsx-only deployments don't need it

    book = xlrd.open_workbook(path, formatting_info=False)
    workbook = openpyxl.Workbook()
    # Remove the default sheet; we recreate sheets by their real names.
    default = workbook.active
    workbook.remove(default)

    for sheet_index in range(book.nsheets):
        source = book.sheet_by_index(sheet_index)
        target = workbook.create_sheet(title=source.name[:31])
        for row in range(source.nrows):
            for col in range(source.ncols):
                cell = source.cell(row, col)
                value = cell.value
                if value == "":
                    continue
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        parsed = xlrd.xldate.xldate_as_datetime(value, book.datemode)
                        value = parsed
                    except Exception:
                        pass
                target.cell(row=row + 1, column=col + 1, value=value)

    return workbook
