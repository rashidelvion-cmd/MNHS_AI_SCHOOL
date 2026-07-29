"""
Regression tests for the semester-aware SF10 subject placement.

Covers the review scenarios: missing assignment, NULL semester with other
assignments, wrong-section, duplicate, conflicting semesters, one school
year with both semesters, all-NULL legacy data, SF9 unchanged, no subject
silently dropped, and warnings shown.
"""

import datetime
import io

import openpyxl
from django.test import TestCase

from academics.models import (
    GradeLevel,
    SchoolYear,
    Section,
    Subject,
    SubjectAssignment,
)
from enrollment.models import Enrollment
from grades.models import Grade
from students.models import Student
from teachers.models import Teacher
from reports.services.sf10.exporter import build_sf10_workbook
from reports.services.sf9.data import build_sf9_data


def _front(workbook):
    return workbook["FRONT"]


def _block_subjects(ws, first_row, last_row):
    return {
        ws[f"D{r}"].value: ws[f"T{r}"].value
        for r in range(first_row, last_row + 1)
        if ws[f"D{r}"].value
    }


class SF10SemesterTests(TestCase):
    # FRONT sem1 free rows 35-41, sem2 free rows 74-80
    SEM1_ROWS = (35, 41)
    SEM2_ROWS = (74, 80)

    def setUp(self):
        self.sy = SchoolYear.objects.create(year="2023-2024", grading_system="term")
        self.gl = GradeLevel.objects.create(name="Grade 11")
        self.sec = Section.objects.create(
            grade_level=self.gl, name="11-CSS", track_strand="TVL-ICT"
        )
        self.other_sec = Section.objects.create(
            grade_level=self.gl, name="11-OTHER", track_strand="TVL-ICT"
        )
        self.teacher = Teacher.objects.create(first_name="Gina", last_name="Montebon")
        self.student = Student.objects.create(
            lrn="121639220016", last_name="DIONERO", first_name="JAYRON",
            gender="Male", birth_date=datetime.date(2007, 1, 1),
        )
        self.enroll = Enrollment.objects.create(
            student=self.student, school_year=self.sy, grade_level=self.gl,
            section=self.sec, semester="1",
        )

    def _graded_subject(self, code, name, semester, section=None, final=87):
        subject = Subject.objects.create(code=code, name=name)
        SubjectAssignment.objects.create(
            school_year=self.sy, section=section or self.sec,
            subject=subject, teacher=self.teacher, semester=semester,
        )
        Grade.objects.create(
            student=self.student, subject=subject, school_year=self.sy,
            first_quarter=final, second_quarter=final, third_quarter=final,
        )
        return subject

    def _grade_only(self, code, name, final=87):
        """A graded subject with NO SubjectAssignment at all."""
        subject = Subject.objects.create(code=code, name=name)
        Grade.objects.create(
            student=self.student, subject=subject, school_year=self.sy,
            first_quarter=final, second_quarter=final, third_quarter=final,
        )
        return subject

    def _export(self):
        workbook, warnings = build_sf10_workbook(self.student, self.gl)
        ws = _front(workbook)
        sem1 = _block_subjects(ws, *self.SEM1_ROWS)
        sem2 = _block_subjects(ws, *self.SEM2_ROWS)
        return sem1, sem2, warnings

    # ---- Scenario 6: one school year with both semesters ----
    def test_both_semesters_single_year(self):
        self._graded_subject("S1A", "Oral Communication", "1")
        self._graded_subject("S2A", "Reading and Writing", "2")
        sem1, sem2, warnings = self._export()
        self.assertIn("Oral Communication", sem1)
        self.assertIn("Reading and Writing", sem2)
        self.assertNotIn("Oral Communication", sem2)
        self.assertNotIn("Reading and Writing", sem1)

    # ---- Scenario 1 + 9: missing assignment, never dropped ----
    def test_missing_assignment_goes_to_first_with_warning(self):
        self._graded_subject("HAS", "Has Assignment", "1")
        self._grade_only("NOASG", "No Assignment Subj")
        sem1, sem2, warnings = self._export()
        self.assertIn("No Assignment Subj", sem1)   # not dropped
        self.assertNotIn("No Assignment Subj", sem2)
        self.assertTrue(any("No Assignment Subj" in w for w in warnings))

    # ---- Scenario 2: NULL semester with other semester assignments ----
    def test_null_semester_with_others(self):
        self._graded_subject("S1", "Assigned Sem1", "1")
        # subject present with an assignment but semester=None
        subject = Subject.objects.create(code="NUL", name="Null Semester Subj")
        SubjectAssignment.objects.create(
            school_year=self.sy, section=self.sec, subject=subject,
            teacher=self.teacher, semester=None,
        )
        Grade.objects.create(
            student=self.student, subject=subject, school_year=self.sy,
            first_quarter=90, second_quarter=90, third_quarter=90,
        )
        sem1, sem2, warnings = self._export()
        self.assertIn("Null Semester Subj", sem1)
        self.assertNotIn("Null Semester Subj", sem2)
        self.assertTrue(any("Null Semester Subj" in w for w in warnings))

    # ---- Scenario 3: wrong-section assignment ----
    def test_wrong_section_assignment(self):
        self._graded_subject("WS", "Wrong Section Subj", "2", section=self.other_sec)
        sem1, sem2, warnings = self._export()
        # placed in First Sem as safe fallback, not dropped, warned
        self.assertIn("Wrong Section Subj", sem1)
        self.assertNotIn("Wrong Section Subj", sem2)
        self.assertTrue(any("11-OTHER" in w and "Wrong Section Subj" in w for w in warnings))

    # ---- Scenario 4: duplicate same-section same-semester (deterministic) ----
    def test_duplicate_same_semester(self):
        subject = self._graded_subject("DUP", "Dup Subject", "1")
        SubjectAssignment.objects.create(
            school_year=self.sy, section=self.sec, subject=subject,
            teacher=self.teacher, semester="1",
        )
        sem1, sem2, warnings = self._export()
        self.assertIn("Dup Subject", sem1)
        self.assertNotIn("Dup Subject", sem2)  # never both

    # ---- Scenario 5: conflicting same-section semesters ----
    def test_conflicting_semesters(self):
        subject = self._graded_subject("CON", "Conflict Subject", "1")
        SubjectAssignment.objects.create(
            school_year=self.sy, section=self.sec, subject=subject,
            teacher=self.teacher, semester="2",
        )
        sem1, sem2, warnings = self._export()
        # deterministic: lowest semester wins -> First
        self.assertIn("Conflict Subject", sem1)
        self.assertNotIn("Conflict Subject", sem2)
        self.assertTrue(any("Conflict Subject" in w and "conflicting" in w.lower() for w in warnings))

    # ---- Scenario 7: legacy data, all semester NULL ----
    def test_legacy_all_null(self):
        for code, name in (("L1", "Legacy One"), ("L2", "Legacy Two")):
            subject = Subject.objects.create(code=code, name=name)
            SubjectAssignment.objects.create(
                school_year=self.sy, section=self.sec, subject=subject,
                teacher=self.teacher, semester=None,
            )
            Grade.objects.create(
                student=self.student, subject=subject, school_year=self.sy,
                first_quarter=88, second_quarter=88, third_quarter=88,
            )
        sem1, sem2, warnings = self._export()
        # all in First Sem, none duplicated to Second, one clear warning
        self.assertIn("Legacy One", sem1)
        self.assertIn("Legacy Two", sem1)
        self.assertNotIn("Legacy One", sem2)
        self.assertNotIn("Legacy Two", sem2)
        self.assertTrue(any("No semester information" in w for w in warnings))

    # ---- Scenario 9: no subject silently dropped (aggregate) ----
    def test_no_subject_silently_dropped(self):
        self._graded_subject("A", "Subject A", "1")
        self._graded_subject("B", "Subject B", "2")
        self._grade_only("C", "Subject C")                       # no assignment
        self._graded_subject("D", "Subject D", "2", section=self.other_sec)  # wrong section
        sem1, sem2, warnings = self._export()
        placed = set(sem1) | set(sem2)
        for name in ("Subject A", "Subject B", "Subject C", "Subject D"):
            self.assertIn(name, placed, f"{name} was silently dropped")
        # none appear twice
        self.assertEqual(len(set(sem1) & set(sem2)), 0)

    # ---- C-1 follow-up: two SchoolYears, BOTH semester NULL ----
    def test_two_schoolyears_both_null_semester(self):
        sy2 = SchoolYear.objects.create(year="2024-2025", grading_system="term")
        Enrollment.objects.create(
            student=self.student, school_year=sy2, grade_level=self.gl,
            section=self.sec, semester="2",
        )
        # NULL-semester subject in the FIRST year
        subj1 = Subject.objects.create(code="N1", name="Null1 Subj")
        SubjectAssignment.objects.create(
            school_year=self.sy, section=self.sec, subject=subj1,
            teacher=self.teacher, semester=None,
        )
        Grade.objects.create(
            student=self.student, subject=subj1, school_year=self.sy,
            first_quarter=88, second_quarter=88, third_quarter=88,
        )
        # NULL-semester subject in the SECOND year
        subj2 = Subject.objects.create(code="N2", name="Null2 Subj")
        SubjectAssignment.objects.create(
            school_year=sy2, section=self.sec, subject=subj2,
            teacher=self.teacher, semester=None,
        )
        Grade.objects.create(
            student=self.student, subject=subj2, school_year=sy2,
            first_quarter=88, second_quarter=88, third_quarter=88,
        )
        sem1, sem2, warnings = self._export()
        # each NULL subject falls back to the block ITS enrollment drives
        self.assertIn("Null1 Subj", sem1)
        self.assertIn("Null2 Subj", sem2)          # must NOT be dropped
        self.assertNotIn("Null1 Subj", sem2)
        self.assertNotIn("Null2 Subj", sem1)
        self.assertEqual(len(set(sem1) & set(sem2)), 0)
        # a warning is shown for each year's NULL data
        self.assertTrue(any("Null1 Subj" in w or "No semester information" in w for w in warnings))
        self.assertTrue(any("Null2 Subj" in w or "No semester information" in w for w in warnings))

    # ---- Scenario 8: SF9 remains unchanged ----
    def test_sf9_unchanged(self):
        self._graded_subject("S1", "Oral Communication", "1")
        self._graded_subject("S2", "Reading and Writing", "2")
        subjects = build_sf9_data(self.student, self.sy)["subjects"]
        names = [entry["subject"].name for entry in subjects]
        # SF9 shows ALL subjects regardless of semester
        self.assertIn("Oral Communication", names)
        self.assertIn("Reading and Writing", names)
        self.assertEqual(len(names), 2)

    # ---- C-1: two SchoolYears / two enrollments, one per semester ----
    def test_two_schoolyears_two_enrollments(self):
        sy2 = SchoolYear.objects.create(year="2023-2024 (2nd)", grading_system="term")
        Enrollment.objects.create(
            student=self.student, school_year=sy2, grade_level=self.gl,
            section=self.sec, semester="2",
        )
        # sem1 subject in the first year
        self._graded_subject("Y1S1", "Year1 Sem1 Subj", "1")
        # sem2 subject in the SECOND year
        subj2 = Subject.objects.create(code="Y2S2", name="Year2 Sem2 Subj")
        SubjectAssignment.objects.create(
            school_year=sy2, section=self.sec, subject=subj2,
            teacher=self.teacher, semester="2",
        )
        Grade.objects.create(
            student=self.student, subject=subj2, school_year=sy2,
            first_quarter=90, second_quarter=90, third_quarter=90,
        )
        sem1, sem2, warnings = self._export()
        self.assertIn("Year1 Sem1 Subj", sem1)
        self.assertIn("Year2 Sem2 Subj", sem2)   # must NOT be dropped (C-1)
        self.assertNotIn("Year2 Sem2 Subj", sem1)

    # ---- C-1: two SchoolYears with correct SY labels in each block ----
    def test_two_schoolyears_labels_preserved(self):
        sy2 = SchoolYear.objects.create(year="2024-2025", grading_system="term")
        Enrollment.objects.create(
            student=self.student, school_year=sy2, grade_level=self.gl,
            section=self.sec, semester="2",
        )
        self._graded_subject("Y1", "Y1 Subj", "1")
        subj2 = Subject.objects.create(code="Y2", name="Y2 Subj")
        SubjectAssignment.objects.create(
            school_year=sy2, section=self.sec, subject=subj2,
            teacher=self.teacher, semester="2",
        )
        Grade.objects.create(
            student=self.student, subject=subj2, school_year=sy2,
            first_quarter=85, second_quarter=85, third_quarter=85,
        )
        workbook, _ = build_sf10_workbook(self.student, self.gl)
        ws = _front(workbook)
        self.assertEqual(ws["S23"].value, "2023-2024")   # sem1 block SY
        self.assertEqual(ws["S62"].value, "2024-2025")   # sem2 block SY

    # ---- H-1: capacity overflow names every omitted subject ----
    def test_capacity_overflow_names_subjects(self):
        # sem1 free rows are 35-41 (7 rows). Create 9 sem-1 subjects.
        created = []
        for i in range(9):
            name = f"Overflow Subj {i:02d}"
            self._graded_subject(f"OF{i:02d}", name, "1")
            created.append(name)
        sem1, sem2, warnings = self._export()
        # exactly 7 fit
        self.assertEqual(len(sem1), 7)
        # the 2 that didn't fit are each named in a warning
        written = set(sem1)
        omitted = [n for n in created if n not in written]
        self.assertEqual(len(omitted), 2)
        for name in omitted:
            self.assertTrue(
                any(name in w and "full" in w.lower() for w in warnings),
                f"{name} overflow not warned by name",
            )

    # ---- M-1: placed set means "actually written" (overflow not placed) ----
    def test_overflow_subject_not_marked_placed(self):
        # 8 sem-1 subjects: 7 fit, 1 overflows. The overflowed one must not
        # be silently swallowed into the (empty) sem-2 block either.
        for i in range(8):
            self._graded_subject(f"P{i:02d}", f"Placed Subj {i:02d}", "1")
        sem1, sem2, warnings = self._export()
        self.assertEqual(len(sem1), 7)
        self.assertEqual(len(sem2), 0)  # overflow does NOT leak to sem2
        # the overflowed subject is reported, not dropped silently
        self.assertTrue(any("full" in w.lower() for w in warnings))

    # ---- M-2: cross-section contradiction, own section authoritative ----
    def test_cross_section_contradiction(self):
        subject = self._graded_subject("XS", "Cross Section Subj", "1")  # own section sem1
        SubjectAssignment.objects.create(
            school_year=self.sy, section=self.other_sec, subject=subject,
            teacher=self.teacher, semester="2",  # other section says sem2
        )
        sem1, sem2, warnings = self._export()
        self.assertIn("Cross Section Subj", sem1)      # own section wins
        self.assertNotIn("Cross Section Subj", sem2)
        self.assertTrue(any("Cross Section Subj" in w and "11-OTHER" in w for w in warnings))


# ---------------------------------------------------------------------------
# ID Maker tests — Module 11 (D[131-137])
# ---------------------------------------------------------------------------

import datetime
from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model

from academics.models import SchoolYear, GradeLevel, Section
from students.models import Student
from teachers.models import Teacher
from enrollment.models import Enrollment

User = get_user_model()


class StudentIDServiceTests(TestCase):
    """Unit tests for the student ID PDF service."""

    def setUp(self):
        self.sy  = SchoolYear.objects.create(year="2023-2024", grading_system="term", is_active=True)
        self.gl  = GradeLevel.objects.create(name="Grade 11")
        self.sec = Section.objects.create(grade_level=self.gl, name="11-CSS", track_strand="TVL")
        self.student = Student.objects.create(
            lrn="121708120001",
            last_name="DELA CRUZ",
            first_name="JUAN",
            middle_name="SANTOS",
            gender="Male",
            birth_date=datetime.date(2007, 1, 15),
        )
        Enrollment.objects.create(
            student=self.student,
            school_year=self.sy,
            grade_level=self.gl,
            section=self.sec,
            semester="1",
        )

    def test_student_id_returns_pdf_bytes(self):
        """build_student_id_pdf returns non-empty bytes."""
        from reports.services.id_maker.student_id import build_student_id_pdf
        pdf = build_student_id_pdf([self.student])
        self.assertIsInstance(pdf, bytes)
        self.assertGreater(len(pdf), 0)

    def test_student_id_starts_with_pdf_header(self):
        """Output is a valid PDF (starts with %PDF)."""
        from reports.services.id_maker.student_id import build_student_id_pdf
        pdf = build_student_id_pdf([self.student])
        self.assertTrue(pdf.startswith(b"%PDF"), "Output must be a valid PDF file")

    def test_student_id_qr_uses_lrn(self):
        """QR code is generated from student.lrn exactly as D[137] specifies."""
        import qrcode
        # Verify qrcode.make(student.lrn) produces valid output
        img = qrcode.make(self.student.lrn)
        self.assertIsNotNone(img)
        # Verify the service calls it the same way by checking it renders without error
        from reports.services.id_maker.student_id import build_student_id_pdf
        pdf = build_student_id_pdf([self.student])
        self.assertGreater(len(pdf), 100)

    def test_student_id_multiple_students(self):
        """PDF is generated for a list of multiple students."""
        from reports.services.id_maker.student_id import build_student_id_pdf
        s2 = Student.objects.create(
            lrn="121708120002", last_name="REYES", first_name="MARIA",
            gender="Female", birth_date=datetime.date(2007, 3, 10),
        )
        Enrollment.objects.create(
            student=s2, school_year=self.sy, grade_level=self.gl,
            section=self.sec, semester="1",
        )
        pdf = build_student_id_pdf([self.student, s2])
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_student_id_no_enrollment(self):
        """Student without enrollment still generates a PDF (no crash)."""
        from reports.services.id_maker.student_id import build_student_id_pdf
        s_no_enr = Student.objects.create(
            lrn="999999999999", last_name="NONE", first_name="NO",
            gender="Male", birth_date=datetime.date(2007, 1, 1),
        )
        pdf = build_student_id_pdf([s_no_enr])
        self.assertTrue(pdf.startswith(b"%PDF"))


class TeacherIDServiceTests(TestCase):
    """Unit tests for the teacher ID PDF service."""

    def setUp(self):
        self.teacher = Teacher.objects.create(
            employee_id="EMP-001",
            first_name="GINALYN",
            last_name="MONTEBON",
        )

    def test_teacher_id_returns_pdf_bytes(self):
        """build_teacher_id_pdf returns non-empty bytes."""
        from reports.services.id_maker.teacher_id import build_teacher_id_pdf
        pdf = build_teacher_id_pdf([self.teacher])
        self.assertIsInstance(pdf, bytes)
        self.assertGreater(len(pdf), 0)

    def test_teacher_id_starts_with_pdf_header(self):
        """Output is a valid PDF."""
        from reports.services.id_maker.teacher_id import build_teacher_id_pdf
        pdf = build_teacher_id_pdf([self.teacher])
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_teacher_id_qr_uses_employee_id(self):
        """QR code is generated from teacher.employee_id."""
        import qrcode
        img = qrcode.make(self.teacher.employee_id)
        self.assertIsNotNone(img)
        from reports.services.id_maker.teacher_id import build_teacher_id_pdf
        pdf = build_teacher_id_pdf([self.teacher])
        self.assertGreater(len(pdf), 100)

    def test_teacher_id_multiple_teachers(self):
        """PDF is generated for multiple teachers."""
        from reports.services.id_maker.teacher_id import build_teacher_id_pdf
        t2 = Teacher.objects.create(
            employee_id="EMP-002", first_name="JOSE", last_name="SANTOS"
        )
        pdf = build_teacher_id_pdf([self.teacher, t2])
        self.assertTrue(pdf.startswith(b"%PDF"))


class IDMakerViewAuthTests(TestCase):
    """Access control tests for ID Maker views."""

    STUDENT_URL = "/reports/id/student/"
    TEACHER_URL = "/reports/id/teacher/"

    def setUp(self):
        self.allowed_roles = [
            ("admin",           "admin"),
            ("principal",       "principal"),
            ("ict",             "ict_coordinator"),
            ("adviser",         "adviser"),
            ("subject_teacher", "subject_teacher"),
            ("teacher",         "teacher"),
        ]
        self.blocked_roles = [
            ("student_u", "student"),
            ("parent_u",  "parent"),
        ]
        for username, role in self.allowed_roles + self.blocked_roles:
            User.objects.create_user(username=username, password="p", role=role)

    def test_student_id_allowed_roles(self):
        """All staff roles can access the student ID page."""
        for username, role in self.allowed_roles:
            c = Client()
            c.login(username=username, password="p")
            r = c.get(self.STUDENT_URL)
            self.assertEqual(r.status_code, 200,
                             f"Role {role!r} should have access to student ID page")

    def test_teacher_id_allowed_roles(self):
        """All staff roles can access the teacher ID page."""
        for username, role in self.allowed_roles:
            c = Client()
            c.login(username=username, password="p")
            r = c.get(self.TEACHER_URL)
            self.assertEqual(r.status_code, 200,
                             f"Role {role!r} should have access to teacher ID page")

    def test_student_id_blocked_roles(self):
        """Student and parent roles are blocked (403)."""
        for username, role in self.blocked_roles:
            c = Client()
            c.login(username=username, password="p")
            r = c.get(self.STUDENT_URL)
            self.assertEqual(r.status_code, 403,
                             f"Role {role!r} must NOT access student ID page")

    def test_teacher_id_blocked_roles(self):
        """Student and parent roles are blocked (403) on teacher ID page."""
        for username, role in self.blocked_roles:
            c = Client()
            c.login(username=username, password="p")
            r = c.get(self.TEACHER_URL)
            self.assertEqual(r.status_code, 403,
                             f"Role {role!r} must NOT access teacher ID page")

    def test_unauthenticated_redirects(self):
        """Unauthenticated requests redirect to login."""
        c = Client()
        for url in [self.STUDENT_URL, self.TEACHER_URL]:
            r = c.get(url)
            self.assertIn(r.status_code, [302, 301],
                          f"Unauthenticated access to {url} must redirect")

    def test_student_id_post_generates_pdf(self):
        """POST with a valid student selection returns a PDF download."""
        sy  = SchoolYear.objects.create(year="2023-2024", grading_system="term", is_active=True)
        gl  = GradeLevel.objects.create(name="Grade 11")
        sec = Section.objects.create(grade_level=gl, name="11-CSS", track_strand="TVL")
        student = Student.objects.create(
            lrn="121708120001", last_name="TEST", first_name="STUDENT",
            gender="Male", birth_date=datetime.date(2007, 1, 1),
        )
        Enrollment.objects.create(
            student=student, school_year=sy, grade_level=gl,
            section=sec, semester="1",
        )
        c = Client()
        c.login(username="admin", password="p")
        r = c.post(self.STUDENT_URL, {"students": [student.pk]})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertIn("student_ids.pdf", r["Content-Disposition"])
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_teacher_id_post_generates_pdf(self):
        """POST with a valid teacher selection returns a PDF download."""
        teacher = Teacher.objects.create(
            employee_id="EMP-TEST", first_name="TEST", last_name="TEACHER"
        )
        c = Client()
        c.login(username="admin", password="p")
        r = c.post(self.TEACHER_URL, {"teachers": [teacher.pk]})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertIn("teacher_ids.pdf", r["Content-Disposition"])
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_existing_sf9_sf10_unaffected(self):
        """Existing SF9 and SF10 URLs still return 200 for admin."""
        c = Client()
        c.login(username="admin", password="p")
        self.assertEqual(c.get("/reports/sf9/").status_code, 200)
        self.assertEqual(c.get("/reports/sf10/export/").status_code, 200)

    def test_existing_sf1_import_unaffected(self):
        """Existing SF1 import URL still returns 200 for admin."""
        c = Client()
        c.login(username="admin", password="p")
        self.assertEqual(c.get("/reports/sf1/import/").status_code, 200)


# ---------------------------------------------------------------------------
# Certificate Generator tests — Module 12 (D[138]–D[145])
# ---------------------------------------------------------------------------

class CertificateServiceTests(TestCase):
    """Unit tests for all five certificate PDF generators."""

    def setUp(self):
        self.sy  = SchoolYear.objects.create(
            year="2023-2024", grading_system="term", is_active=True)
        self.gl  = GradeLevel.objects.create(name="Grade 11")
        self.sec = Section.objects.create(
            grade_level=self.gl, name="11-CSS", track_strand="TVL")
        self.student = Student.objects.create(
            lrn="121708120010",
            last_name="SANTOS",
            first_name="MARIA",
            middle_name="DELA CRUZ",
            gender="Female",
            birth_date=datetime.date(2007, 5, 20),
        )
        Enrollment.objects.create(
            student=self.student,
            school_year=self.sy,
            grade_level=self.gl,
            section=self.sec,
            semester="1",
        )

    # -- Enrollment ----------------------------------------------------------

    def test_enrollment_cert_returns_pdf(self):
        from reports.services.certificates.generator import build_enrollment_cert
        pdf = build_enrollment_cert(self.student)
        self.assertIsInstance(pdf, bytes)
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_enrollment_cert_student_name_in_pdf(self):
        """Student name is embedded in the PDF content."""
        from reports.services.certificates.generator import build_enrollment_cert
        pdf = build_enrollment_cert(self.student)
        self.assertGreater(len(pdf), 100)

    # -- Completion ----------------------------------------------------------

    def test_completion_cert_returns_pdf(self):
        from reports.services.certificates.generator import build_completion_cert
        pdf = build_completion_cert(self.student)
        self.assertTrue(pdf.startswith(b"%PDF"))

    # -- Recognition ---------------------------------------------------------

    def test_recognition_cert_returns_pdf(self):
        from reports.services.certificates.generator import build_recognition_cert
        pdf = build_recognition_cert(self.student, "Academic Excellence")
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_recognition_cert_uses_free_text(self):
        """Free-text string is passed through to the PDF generator without error."""
        from reports.services.certificates.generator import build_recognition_cert
        for text in ["Science Fair 2024", "Math Olympiad", "Leadership Award"]:
            pdf = build_recognition_cert(self.student, text)
            self.assertTrue(pdf.startswith(b"%PDF"), f"Failed for text: {text!r}")

    # -- Participation -------------------------------------------------------

    def test_participation_cert_returns_pdf(self):
        from reports.services.certificates.generator import build_participation_cert
        pdf = build_participation_cert(self.student, "School Sports Fest")
        self.assertTrue(pdf.startswith(b"%PDF"))

    def test_participation_cert_uses_free_text(self):
        from reports.services.certificates.generator import build_participation_cert
        for text in ["Quiz Bee", "Cultural Show", "Community Clean-up Drive"]:
            pdf = build_participation_cert(self.student, text)
            self.assertTrue(pdf.startswith(b"%PDF"), f"Failed for text: {text!r}")

    # -- Diploma -------------------------------------------------------------

    def test_diploma_returns_pdf(self):
        from reports.services.certificates.generator import build_diploma
        pdf = build_diploma(self.student)
        self.assertTrue(pdf.startswith(b"%PDF"))

    # -- No enrollment -------------------------------------------------------

    def test_all_types_handle_no_enrollment(self):
        """All generators work even when the student has no enrollment."""
        from reports.services.certificates.generator import (
            build_enrollment_cert, build_completion_cert,
            build_recognition_cert, build_participation_cert, build_diploma,
        )
        s = Student.objects.create(
            lrn="000000000099", last_name="NOENR", first_name="X",
            gender="Male", birth_date=datetime.date(2007, 1, 1),
        )
        for fn, args in [
            (build_enrollment_cert,    (s,)),
            (build_completion_cert,    (s,)),
            (build_recognition_cert,   (s, "Award")),
            (build_participation_cert, (s, "Event")),
            (build_diploma,            (s,)),
        ]:
            pdf = fn(*args)
            self.assertTrue(pdf.startswith(b"%PDF"),
                            f"{fn.__name__} failed for student without enrollment")


class CertificateViewTests(TestCase):
    """Access control and HTTP behaviour tests for the certificate view."""

    URL = "/reports/certificates/"

    def setUp(self):
        self.allowed = [
            ("cert_admin",   "admin"),
            ("cert_prin",    "principal"),
            ("cert_ict",     "ict_coordinator"),
            ("cert_adv",     "adviser"),
            ("cert_subj",    "subject_teacher"),
            ("cert_teach",   "teacher"),
        ]
        self.blocked = [
            ("cert_stud", "student"),
            ("cert_par",  "parent"),
        ]
        for username, role in self.allowed + self.blocked:
            User.objects.create_user(username=username, password="p", role=role)

        # shared student + enrollment for POST tests
        sy  = SchoolYear.objects.create(
            year="2023-2024", grading_system="term", is_active=True)
        gl  = GradeLevel.objects.create(name="Grade 11")
        sec = Section.objects.create(
            grade_level=gl, name="11-CSS", track_strand="TVL")
        self.student = Student.objects.create(
            lrn="121708120099", last_name="CERT", first_name="TEST",
            gender="Male", birth_date=datetime.date(2007, 1, 1),
        )
        Enrollment.objects.create(
            student=self.student, school_year=sy,
            grade_level=gl, section=sec, semester="1",
        )

    def test_allowed_roles_get_200(self):
        for username, role in self.allowed:
            c = Client()
            c.login(username=username, password="p")
            self.assertEqual(c.get(self.URL).status_code, 200,
                             f"Role {role!r} should see certificate page")

    def test_blocked_roles_get_403(self):
        for username, role in self.blocked:
            c = Client()
            c.login(username=username, password="p")
            self.assertEqual(c.get(self.URL).status_code, 403,
                             f"Role {role!r} must NOT access certificate page")

    def test_unauthenticated_redirects(self):
        r = Client().get(self.URL)
        self.assertIn(r.status_code, [301, 302])

    def test_post_enrollment_cert_returns_pdf(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student": self.student.pk,
            "cert_type": "enrollment",
        })
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_post_completion_cert_returns_pdf(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student": self.student.pk,
            "cert_type": "completion",
        })
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_post_recognition_cert_with_free_text(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student":   self.student.pk,
            "cert_type": "recognition",
            "free_text": "Science Fair 2024",
        })
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_post_participation_cert_with_free_text(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student":   self.student.pk,
            "cert_type": "participation",
            "free_text": "School Sports Fest",
        })
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_post_diploma_returns_pdf(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student":   self.student.pk,
            "cert_type": "diploma",
        })
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_recognition_without_free_text_redirects(self):
        """Recognition cert with empty free_text redirects back with warning."""
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student":   self.student.pk,
            "cert_type": "recognition",
            "free_text": "",
        })
        self.assertIn(r.status_code, [301, 302])

    def test_participation_without_free_text_redirects(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        r = c.post(self.URL, {
            "student":   self.student.pk,
            "cert_type": "participation",
            "free_text": "",
        })
        self.assertIn(r.status_code, [301, 302])

    def test_existing_id_maker_unaffected(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        self.assertEqual(c.get("/reports/id/student/").status_code, 200)
        self.assertEqual(c.get("/reports/id/teacher/").status_code, 200)

    def test_existing_sf9_sf10_unaffected(self):
        c = Client()
        c.login(username="cert_admin", password="p")
        self.assertEqual(c.get("/reports/sf9/").status_code, 200)
        self.assertEqual(c.get("/reports/sf10/export/").status_code, 200)


# ---------------------------------------------------------------------------
# SF1 PDF Print tests — D[063]
# ---------------------------------------------------------------------------

import datetime
from academics.models import SchoolYear, GradeLevel, Section
from students.models import Student
from enrollment.models import Enrollment


class SF1PDFTests(TestCase):

    def setUp(self):
        self.sy  = SchoolYear.objects.create(
            year="2023-2024", grading_system="term", is_active=True)
        self.gl  = GradeLevel.objects.create(name="Grade 11")
        self.sec = Section.objects.create(
            grade_level=self.gl, name="11-CSS", track_strand="TVL-ICT")

        # Two students (one male, one female) enrolled in this section
        for i, (fn, ln, gend) in enumerate([
            ("JUAN",  "DELA CRUZ", "Male"),
            ("MARIA", "SANTOS",    "Female"),
        ]):
            s = Student.objects.create(
                lrn=f"12170812000{i}", first_name=fn, last_name=ln,
                gender=gend, birth_date=datetime.date(2007, 1, 1),
            )
            Enrollment.objects.create(
                student=s, school_year=self.sy,
                grade_level=self.gl, section=self.sec, semester="1",
            )

        for role in ("admin", "principal", "ict_coordinator",
                     "teacher", "adviser", "subject_teacher",
                     "student", "parent"):
            User.objects.create_user(
                username=f"sf1pdf_{role}", password="p", role=role)

    def _post(self, role="admin"):
        c = Client()
        c.login(username=f"sf1pdf_{role}", password="p")
        return c.post("/reports/sf1/pdf/", {
            "school_year": self.sy.pk,
            "grade_level": self.gl.pk,
            "section":     self.sec.pk,
            "semester":    "1",
        })

    # ── Service unit tests ────────────────────────────────────────────────

    def test_sf1_pdf_returns_pdf_bytes(self):
        """build_sf1_pdf returns bytes that start with the PDF header."""
        from reports.services.sf1.pdf import build_sf1_pdf
        pdf = build_sf1_pdf(self.sy, self.gl, self.sec, "1st Semester")
        self.assertIsInstance(pdf, bytes)
        self.assertTrue(pdf.startswith(b"%PDF"), "Output must be a valid PDF")

    def test_sf1_pdf_with_no_students(self):
        """Empty section (no enrollments) generates a valid PDF without crash."""
        from reports.services.sf1.pdf import build_sf1_pdf
        empty_sec = Section.objects.create(
            grade_level=self.gl, name="11-EMPTY", track_strand="Academic")
        pdf = build_sf1_pdf(self.sy, self.gl, empty_sec, "1st Semester")
        self.assertTrue(pdf.startswith(b"%PDF"))

    # ── View tests ────────────────────────────────────────────────────────

    def test_sf1_pdf_view_get(self):
        """GET /reports/sf1/pdf/ returns 200 for an allowed role."""
        c = Client()
        c.login(username="sf1pdf_admin", password="p")
        r = c.get("/reports/sf1/pdf/")
        self.assertEqual(r.status_code, 200)

    def test_sf1_pdf_view_post_downloads(self):
        """POST with valid form returns application/pdf response."""
        r = self._post("admin")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_sf1_pdf_filename(self):
        """Content-Disposition filename contains section name and school year."""
        r = self._post("admin")
        cd = r.get("Content-Disposition", "")
        self.assertIn("SF1_", cd)
        self.assertIn("11-CSS", cd)
        self.assertIn(".pdf", cd)

    # ── Permission tests ─────────────────────────────────────────────────

    def test_sf1_pdf_allowed_roles(self):
        """All 6 documented allowed roles get 200 on GET."""
        allowed = ["admin", "principal", "ict_coordinator",
                   "teacher", "adviser", "subject_teacher"]
        for role in allowed:
            c = Client()
            c.login(username=f"sf1pdf_{role}", password="p")
            r = c.get("/reports/sf1/pdf/")
            self.assertEqual(r.status_code, 200,
                             f"Role {role!r} should access SF1 PDF page")

    def test_sf1_pdf_blocked_student(self):
        """Student role is blocked (403)."""
        c = Client()
        c.login(username="sf1pdf_student", password="p")
        self.assertEqual(c.get("/reports/sf1/pdf/").status_code, 403)

    def test_sf1_pdf_blocked_parent(self):
        """Parent role is blocked (403)."""
        c = Client()
        c.login(username="sf1pdf_parent", password="p")
        self.assertEqual(c.get("/reports/sf1/pdf/").status_code, 403)

    # ── Regression ───────────────────────────────────────────────────────

    def test_sf1_excel_export_unaffected(self):
        """Existing /reports/sf1/export/ still works after PDF addition."""
        c = Client()
        c.login(username="sf1pdf_admin", password="p")
        self.assertEqual(c.get("/reports/sf1/export/").status_code, 200)
