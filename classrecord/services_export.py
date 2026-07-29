"""
ECR Export service — M-7Jul (Messenger requirement).

Client requirement (verbatim, 7 Jul, Mhar):
    "Subject Teacher upload SF1 then data is automatic to SF2 then open
     E-ClassRecord inputs grade then send the softcopy to the Adviser or
     else generate E-ClassRecord with grades"

Provides two export formats:
    build_ecr_excel(assignment, quarter) → bytes  (.xlsx)
    build_ecr_pdf(assignment, quarter)   → bytes  (.pdf)

Data source: same ScoreItem / Assessment / compute_grades queries used by
ecr_grid. No new model, no migration, no schema change.
"""

from __future__ import annotations

import datetime
from io import BytesIO

from django.conf import settings

from .models import ScoreItem, Assessment, COMPONENT_CHOICES
from .services import compute_grades, get_weights
from enrollment.models import Enrollment

# Inline helpers (mirrors classrecord/views.py — avoids circular import)
_COMPONENT_ORDER = {"WW": 0, "PT": 1, "QA": 2}


def _ordered_assessments(assignment, quarter):
    return sorted(
        Assessment.objects.filter(subject_assignment=assignment, quarter=quarter),
        key=lambda a: (_COMPONENT_ORDER.get(a.component, 9), a.order, a.pk),
    )


def _ecr_students(assignment):
    enrollments = (
        Enrollment.objects
        .filter(school_year=assignment.school_year, section=assignment.section)
        .select_related("student")
        .order_by("student__gender", "student__last_name", "student__first_name")
    )
    return [e.student for e in enrollments]


def _collect_data(assignment, quarter: int) -> dict:
    """
    Gather all ECR data for one assignment + quarter.
    Returns a dict consumed by both Excel and PDF builders.
    """
    school_year = assignment.school_year
    subject     = assignment.subject
    students    = _ecr_students(assignment)
    assessments = _ordered_assessments(assignment, quarter)

    scores = {
        (item.student_id, item.assessment_id): item.raw_score
        for item in ScoreItem.objects.filter(
            assessment__in=assessments,
            student__in=students,
        )
    }

    rows = []
    for student in students:
        initial_grade, term_grade, _ = compute_grades(
            student, subject, school_year, quarter
        )
        rows.append({
            "student":       student,
            "cells": [
                scores.get((student.pk, a.pk), "")
                for a in assessments
            ],
            "initial_grade": initial_grade,
            "term_grade":    term_grade,
        })

    weights = get_weights(subject)

    return {
        "assignment":   assignment,
        "subject":      subject,
        "section":      assignment.section,
        "school_year":  school_year,
        "teacher":      assignment.teacher,
        "quarter":      quarter,
        "period_label": school_year.period_label(quarter),
        "assessments":  assessments,
        "weights":      weights,
        "rows":         rows,
        "component_labels": dict(COMPONENT_CHOICES),
    }


# ── Excel export ──────────────────────────────────────────────────────────────

def build_ecr_excel(assignment, quarter: int) -> bytes:
    """
    Generate the ECR as an .xlsx workbook and return raw bytes.
    Columns: # | Student Name | [assessment columns] | Initial Grade | Term Grade
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data   = _collect_data(assignment, quarter)
    wb     = openpyxl.Workbook()
    ws     = wb.active
    ws.title = "ECR"

    MAROON = "7A1020"
    GREEN  = "1A5C2E"
    LIGHT  = "F0F0F0"

    def hdr(text, bold=True, bg=None, fg="FFFFFF"):
        """Return a styled header value cell."""
        cell_val = text
        return cell_val

    # ── Title block ──────────────────────────────────────────────────────────
    school_name = getattr(settings, "SCHOOL_NAME", "Mayorga National High School")
    ws.merge_cells("A1:Z1")
    ws["A1"] = "E-CLASS RECORD"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=MAROON)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    meta = [
        ("School",      school_name),
        ("Subject",     str(data["subject"])),
        ("Section",     str(data["section"])),
        ("School Year", str(data["school_year"])),
        ("Period",      data["period_label"]),
        ("Teacher",     str(data["teacher"])),
        ("Weights",     f"WW:{data['weights']['WW']}%  PT:{data['weights']['PT']}%  QA:{data['weights']['QA']}%"),
    ]
    for offset, (label, value) in enumerate(meta, start=2):
        ws.cell(row=offset, column=1, value=label).font = Font(bold=True)
        ws.cell(row=offset, column=2, value=value)

    header_row = len(meta) + 3   # leave a blank row

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["#", "Student Name"]
    for a in data["assessments"]:
        hdr_text = f"{a.component} {a.label}\n(/{a.highest_score})"
        if a.competency:
            hdr_text += f"\n{a.competency[:60]}"
        headers.append(hdr_text)
    headers += ["Initial Grade", "Term Grade"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.row_dimensions[header_row].height = 30

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Data rows ────────────────────────────────────────────────────────────
    for i, row in enumerate(data["rows"], start=1):
        r = header_row + i
        data_row = (
            [i, f"{row['student'].last_name}, {row['student'].first_name}"]
            + [c if c != "" else "" for c in row["cells"]]
            + [
                float(row["initial_grade"]) if row["initial_grade"] is not None else "",
                float(row["term_grade"])    if row["term_grade"]    is not None else "",
            ]
        )
        fill = PatternFill("solid", fgColor=LIGHT) if i % 2 == 0 else None
        for col, val in enumerate(data_row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")
            if col == 2:
                cell.alignment = Alignment(horizontal="left")
            if fill:
                cell.fill = fill

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    for col_idx in range(3, 3 + len(data["assessments"])):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 10
    last_two = [
        openpyxl.utils.get_column_letter(3 + len(data["assessments"])),
        openpyxl.utils.get_column_letter(4 + len(data["assessments"])),
    ]
    for cl in last_two:
        ws.column_dimensions[cl].width = 14

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── PDF export ────────────────────────────────────────────────────────────────

def build_ecr_pdf(assignment, quarter: int) -> bytes:
    """
    Generate the ECR as a landscape-letter PDF and return raw bytes.
    Pattern follows reports/services/sf1/pdf.py exactly.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    MAROON  = colors.HexColor("#7A1020")
    GREEN   = colors.HexColor("#1A5C2E")
    LIGHT   = colors.HexColor("#F0F0F0")
    WHITE   = colors.white
    DARK    = colors.HexColor("#1A1A1A")

    data = _collect_data(assignment, quarter)

    def _s(size, bold=False, color=None, align=TA_CENTER):
        return ParagraphStyle(
            "_", fontSize=size,
            fontName="Helvetica-Bold" if bold else "Helvetica",
            textColor=color or DARK, alignment=align, leading=size * 1.35,
        )

    buffer = BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=0.4 * inch, bottomMargin=0.4 * inch,
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        title=f"ECR {data['subject']} {data['section']} {data['school_year']}",
    )

    elements = []
    school_name = getattr(settings, "SCHOOL_NAME", "Mayorga National High School")
    school_addr = getattr(settings, "SCHOOL_ADDRESS", "Mayorga, Leyte")

    # Header
    elements.append(Paragraph("Republic of the Philippines · Department of Education", _s(7)))
    elements.append(Paragraph(f"<b>{school_name}</b>", _s(10, bold=True, color=MAROON)))
    elements.append(Paragraph(school_addr, _s(7)))
    elements.append(Spacer(1, 3))
    elements.append(Paragraph("<b>ELECTRONIC CLASS RECORD (ECR)</b>", _s(10, bold=True)))
    elements.append(Spacer(1, 3))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=MAROON))
    elements.append(Spacer(1, 3))

    # Meta row
    meta_data = [[
        Paragraph(f"<b>Subject:</b> {data['subject']}", _s(7, align=TA_LEFT)),
        Paragraph(f"<b>Section:</b> {data['section']}", _s(7, align=TA_LEFT)),
        Paragraph(f"<b>School Year:</b> {data['school_year']}", _s(7, align=TA_LEFT)),
        Paragraph(f"<b>Period:</b> {data['period_label']}", _s(7, align=TA_LEFT)),
        Paragraph(f"<b>Teacher:</b> {data['teacher']}", _s(7, align=TA_LEFT)),
        Paragraph(
            f"<b>Weights:</b> WW:{data['weights']['WW']}%  "
            f"PT:{data['weights']['PT']}%  QA:{data['weights']['QA']}%",
            _s(7, align=TA_LEFT),
        ),
    ]]
    PAGE_W = landscape(letter)[0]
    CONTENT_W = PAGE_W - 0.8 * inch
    meta_tbl = Table(meta_data, colWidths=[CONTENT_W / 6] * 6)
    meta_tbl.setStyle(TableStyle([
        ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 5))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    elements.append(Spacer(1, 4))

    # Table header
    col_headers = ["#", "Student Name"]
    for a in data["assessments"]:
        col_hdr = f"{a.component}\n{a.label}\n(/{a.highest_score})"
        if a.competency:
            col_hdr += f"\n{a.competency[:60]}"
        col_headers.append(col_hdr)
    col_headers += ["Initial\nGrade", "Term\nGrade"]

    hdr_row  = [Paragraph(h, _s(5.5, bold=True, color=WHITE)) for h in col_headers]
    tbl_rows = [hdr_row]

    for i, row in enumerate(data["rows"], start=1):
        student_name = (
            f"{row['student'].last_name}, {row['student'].first_name}"
        )
        cells = [
            str(c) if c != "" else "—"
            for c in row["cells"]
        ]
        ig = str(row["initial_grade"]) if row["initial_grade"] is not None else "—"
        tg = str(row["term_grade"])    if row["term_grade"]    is not None else "—"

        data_row = (
            [Paragraph(str(i), _s(5.5)),
             Paragraph(student_name, _s(5.5, align=TA_LEFT))]
            + [Paragraph(c, _s(5.5)) for c in cells]
            + [Paragraph(ig, _s(5.5)),
               Paragraph(tg, _s(5.5, bold=True))]
        )
        tbl_rows.append(data_row)

    # Column widths
    n_assessments = len(data["assessments"])
    fixed_w   = 0.22 * inch + 1.5 * inch  # # + name
    grade_w   = 0.55 * inch               # initial + term
    remaining = CONTENT_W - fixed_w - grade_w * 2
    assess_w  = (remaining / n_assessments) if n_assessments else 0.5 * inch

    col_widths = (
        [0.22 * inch, 1.5 * inch]
        + [assess_w] * n_assessments
        + [0.55 * inch, 0.55 * inch]
    )

    grid_tbl = Table(tbl_rows, colWidths=col_widths, repeatRows=1)
    grid_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), GREEN),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1), 5.5),
        ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
        ("ALIGN",          (1, 1), (1, -1),  "LEFT"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
        ("GRID",           (0, 0), (-1, -1), 0.25, colors.grey),
        ("TOPPADDING",     (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 1),
        ("LEFTPADDING",    (0, 0), (-1, -1), 2),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 2),
    ]))
    elements.append(grid_tbl)

    elements.append(Spacer(1, 8))
    elements.append(
        Paragraph(
            f"Generated: {datetime.date.today().strftime('%B %d, %Y')}  ·  {school_name}",
            _s(6, color=colors.grey),
        )
    )

    doc.build(elements)
    return buffer.getvalue()
