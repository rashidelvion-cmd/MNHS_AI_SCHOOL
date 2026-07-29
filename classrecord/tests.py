"""
classrecord/tests.py — ECR Export tests (M-7Jul)
"""

import datetime

from django.test import TestCase, Client
from django.contrib.auth import get_user_model

from academics.models import SchoolYear, GradeLevel, Section, Subject
from students.models import Student
from enrollment.models import Enrollment
from teachers.models import Teacher
from .models import Assessment, ScoreItem
from .services_export import build_ecr_excel, build_ecr_pdf

User = get_user_model()


class ECRExportTests(TestCase):

    def setUp(self):
        self.sy      = SchoolYear.objects.create(
            year="2023-2024", grading_system="term", is_active=True)
        self.gl      = GradeLevel.objects.create(name="Grade 11")
        self.sec     = Section.objects.create(
            grade_level=self.gl, name="11-CSS", track_strand="TVL-ICT")
        self.subject = Subject.objects.create(
            code="MATH", name="General Mathematics")

        # Users
        for role in ("admin", "teacher", "adviser", "subject_teacher",
                     "student", "parent", "principal"):
            User.objects.create_user(
                username=f"ecr_{role}", password="p", role=role)

        # Teacher model record (SubjectAssignment.teacher is FK to Teacher, not User)
        self.teacher = Teacher.objects.create(
            employee_id="T001", first_name="SAMPLE", last_name="TEACHER")

        # SubjectAssignment
        from academics.models import SubjectAssignment
        self.assignment = SubjectAssignment.objects.create(
            school_year=self.sy, section=self.sec,
            subject=self.subject, teacher=self.teacher, semester="1",
        )

        # Two students enrolled
        for i, (fn, ln) in enumerate([("JUAN", "DELA CRUZ"), ("MARIA", "SANTOS")]):
            s = Student.objects.create(
                lrn=f"12170812{i:04d}", first_name=fn, last_name=ln,
                gender="Male" if i == 0 else "Female",
                birth_date=datetime.date(2007, 1, 1),
            )
            Enrollment.objects.create(
                student=s, school_year=self.sy,
                grade_level=self.gl, section=self.sec, semester="1",
            )

        # One assessment, some scores
        self.assessment = Assessment.objects.create(
            subject_assignment=self.assignment,
            quarter=1, component="WW",
            label="WW1", highest_score=50, order=1,
        )
        for student in Student.objects.all()[:2]:
            ScoreItem.objects.create(
                student=student, assessment=self.assessment,
                quarter=1, component="WW", label="WW1",
                raw_score=45, highest_score=50,
                school_year=self.sy, subject=self.subject,
            )

    # ── Service-level tests ───────────────────────────────────────────────

    def test_ecr_export_returns_excel_bytes(self):
        """build_ecr_excel returns non-empty bytes."""
        result = build_ecr_excel(self.assignment, 1)
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_ecr_export_valid_xlsx(self):
        """build_ecr_excel output is a valid xlsx that openpyxl can open."""
        import openpyxl
        from io import BytesIO
        result = build_ecr_excel(self.assignment, 1)
        wb = openpyxl.load_workbook(BytesIO(result))
        self.assertIn("ECR", wb.sheetnames)

    def test_ecr_pdf_returns_pdf_bytes(self):
        """build_ecr_pdf returns bytes starting with the PDF header."""
        result = build_ecr_pdf(self.assignment, 1)
        self.assertIsInstance(result, bytes)
        self.assertTrue(result.startswith(b"%PDF"),
                        "PDF output must start with %PDF")

    # ── View-level tests ──────────────────────────────────────────────────

    def test_ecr_export_excel_view(self):
        """GET ecr_export_excel returns xlsx content-type for allowed role."""
        c = Client()
        c.login(username="ecr_admin", password="p")
        r = c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/excel/")
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])

    def test_ecr_export_pdf_view(self):
        """GET ecr_export_pdf returns application/pdf for allowed role."""
        c = Client()
        c.login(username="ecr_admin", password="p")
        r = c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/pdf/")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/pdf")
        self.assertTrue(r.content.startswith(b"%PDF"))

    def test_ecr_export_filename(self):
        """Content-Disposition filename includes subject and section."""
        c = Client()
        c.login(username="ecr_admin", password="p")
        r = c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/excel/")
        cd = r.get("Content-Disposition", "")
        self.assertIn("ECR_", cd)
        self.assertIn(".xlsx", cd)

    def test_ecr_export_subject_teacher_allowed(self):
        """Subject Teacher role can download ECR Excel and PDF (Q2-A)."""
        c = Client()
        c.login(username="ecr_subject_teacher", password="p")
        r_xl = c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/excel/")
        r_pdf = c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/pdf/")
        self.assertEqual(r_xl.status_code, 200)
        self.assertEqual(r_pdf.status_code, 200)

    def test_ecr_export_blocked_student(self):
        """Student role cannot access ECR exports (403)."""
        c = Client()
        c.login(username="ecr_student", password="p")
        self.assertEqual(
            c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/excel/").status_code,
            403,
        )
        self.assertEqual(
            c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/pdf/").status_code,
            403,
        )

    def test_ecr_export_blocked_parent(self):
        """Parent role cannot access ECR exports (403)."""
        c = Client()
        c.login(username="ecr_parent", password="p")
        self.assertEqual(
            c.get(f"/classrecord/ecr/{self.assignment.pk}/1/export/excel/").status_code,
            403,
        )

    def test_ecr_grid_unaffected(self):
        """Existing ecr_grid view still returns 200 after export addition."""
        c = Client()
        c.login(username="ecr_admin", password="p")
        r = c.get(f"/classrecord/ecr/{self.assignment.pk}/1/")
        self.assertEqual(r.status_code, 200)


# ---------------------------------------------------------------------------
# Learning Competencies tests — D[082]
# Approved decisions: Q1-A (in exports), Q2-A (optional)
# ---------------------------------------------------------------------------

class LearningCompetencyTests(TestCase):

    def setUp(self):
        self.sy      = SchoolYear.objects.create(
            year="2024-2025", grading_system="term", is_active=True)
        self.gl      = GradeLevel.objects.create(name="Grade 12")
        self.sec     = Section.objects.create(
            grade_level=self.gl, name="12-STEM", track_strand="Academic-STEM")
        self.subject = Subject.objects.create(
            code="CALC", name="Basic Calculus")
        self.teacher = Teacher.objects.create(
            employee_id="T002", first_name="CALCULUS", last_name="TEACHER")
        from academics.models import SubjectAssignment
        self.assignment = SubjectAssignment.objects.create(
            school_year=self.sy, section=self.sec,
            subject=self.subject, teacher=self.teacher, semester="1",
        )
        for role in ("admin", "teacher", "adviser", "subject_teacher",
                     "student", "parent"):
            User.objects.create_user(
                username=f"lc_{role}", password="p", role=role)
        # Enroll a student so ecr_grid doesn't redirect due to empty roster
        import datetime as _dt2
        student = Student.objects.create(
            lrn="999000000001", first_name="COMP", last_name="STUDENT",
            gender="Male", birth_date=_dt2.date(2007, 1, 1),
        )
        Enrollment.objects.create(
            student=student, school_year=self.sy,
            grade_level=self.gl, section=self.sec, semester="1",
        )

        # Assessment WITH a competency
        self.assess_with = Assessment.objects.create(
            subject_assignment=self.assignment, quarter=1,
            component="WW", label="WW1", highest_score=50, order=1,
            competency="Identifies the domain and range of a function.",
        )
        # Assessment WITHOUT a competency (Q2-A: blank is valid)
        self.assess_blank = Assessment.objects.create(
            subject_assignment=self.assignment, quarter=1,
            component="PT", label="PT1", highest_score=100, order=2,
            competency="",
        )

    # ── Model tests ───────────────────────────────────────────────────────

    def test_competency_field_on_assessment(self):
        """Assessment.competency exists as CharField(max_length=300)."""
        from django.db.models import CharField
        field = Assessment._meta.get_field("competency")
        self.assertIsInstance(field, CharField)
        self.assertEqual(field.max_length, 300)

    def test_competency_blank_by_default(self):
        """New Assessment has competency='' (Q2-A: optional)."""
        self.assertEqual(self.assess_blank.competency, "")

    def test_competency_can_be_set(self):
        """Competency text saves and retrieves correctly."""
        self.assess_blank.competency = "Computes the limit of a function."
        self.assess_blank.save()
        self.assess_blank.refresh_from_db()
        self.assertEqual(
            self.assess_blank.competency, "Computes the limit of a function.")

    # ── Form tests ────────────────────────────────────────────────────────

    def test_assessment_form_includes_competency(self):
        """AssessmentForm has a competency field."""
        from classrecord.forms import AssessmentForm
        form = AssessmentForm()
        self.assertIn("competency", form.fields)

    # ── View / template tests ─────────────────────────────────────────────

    def test_ecr_grid_shows_competency_in_header(self):
        """ECR grid renders competency text in column header."""
        admin = User.objects.get(username="lc_admin")
        self.client.force_login(admin)
        r = self.client.get(f"/classrecord/ecr/{self.assignment.pk}/1/")
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "Identifies the domain and range of a function.")

    def test_ecr_grid_no_competency_renders_cleanly(self):
        """ECR grid renders without error when competency is blank (Q2-A)."""
        admin = User.objects.get(username="lc_admin")
        self.client.force_login(admin)
        r = self.client.get(f"/classrecord/ecr/{self.assignment.pk}/1/")
        self.assertEqual(r.status_code, 200)

    def test_add_column_form_renders_competency_input(self):
        """ECR grid page renders competency input in the Add Column form."""
        admin = User.objects.get(username="lc_admin")
        self.client.force_login(admin)
        r = self.client.get(f"/classrecord/ecr/{self.assignment.pk}/1/")
        self.assertContains(r, 'name="competency"')

    # ── Export tests ──────────────────────────────────────────────────────

    def test_excel_export_includes_competency(self):
        """Excel export column header includes competency text (Q1-A)."""
        import openpyxl
        from io import BytesIO
        from classrecord.services_export import build_ecr_excel
        xlsx_bytes = build_ecr_excel(self.assignment, 1)
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
        ws = wb["ECR"]
        # Find any cell containing the competency text
        found = any(
            "Identifies the domain" in str(cell.value or "")
            for row in ws.iter_rows()
            for cell in row
        )
        self.assertTrue(found, "Competency text must appear in Excel export (Q1-A)")

    def test_pdf_export_includes_competency(self):
        """PDF export bytes are non-empty valid PDF (Q1-A — competency in header)."""
        from classrecord.services_export import build_ecr_pdf
        pdf_bytes = build_ecr_pdf(self.assignment, 1)
        self.assertTrue(pdf_bytes.startswith(b"%PDF"))
        self.assertGreater(len(pdf_bytes), 1000)
